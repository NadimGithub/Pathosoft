import logging
logger = logging.getLogger('pathosoft')

from multiprocessing import context
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages

from tests.models import TestMaster
from .forms import RegisterForm, LoginForm
from .models import CustomUser
from .forms import CustomUserCreationForm , CustomUserChangeForm
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from patient.models import PatientMaster
from doctor.models import Doctor    
from patient.models import PatientMaster
from tests.models import TestMaster as Test
from doctor.models import Doctor
import json
from datetime import date, timedelta, datetime
from django.contrib.auth import update_session_auth_hash

@login_required
def dashboard(request):
    try:
        logger.info(f"Dashboard accessed by user: {request.user.username}")
        
        # Get date range from request or use default (last 7 days)
        end_date = date.today()
        start_date = end_date - timedelta(days=6)  # Default to last 7 days
        
        if request.method == 'GET' and 'start_date' in request.GET and 'end_date' in request.GET:
            try:
                start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
                end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
                
                # Validate date range (max 30 days for performance)
                if (end_date - start_date).days > 30:
                    messages.warning(request, 'Date range cannot exceed 30 days. Showing last 30 days.')
                    start_date = end_date - timedelta(days=29)
                    
                if start_date > end_date:
                    start_date, end_date = end_date, start_date  # Swap if start > end
                    
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid date format submitted: {e}")
                messages.error(request, 'Invalid date format. Using default date range.')
                start_date = end_date - timedelta(days=6)
    
        # Generate date range for the chart
        date_list = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
        
        # Get data for the selected date range
        date_labels = [d.strftime("%Y-%m-%d") for d in date_list]
        daily_counts = [PatientMaster.objects.filter(reporting_date=d, status=True).count() for d in date_list]
        
        # Calculate summary statistics
        total_reports = sum(daily_counts)
        avg_daily_reports = round(total_reports / len(daily_counts), 1) if daily_counts else 0
        
        # Get other dashboard data
        test_count = TestMaster.objects.filter(status=True).count()
        patient_count = PatientMaster.objects.filter(status=True).count()
        doctor_count = Doctor.objects.filter(status=True).count()
        pending_reports = PatientMaster.objects.filter(report_status="Pending", status=True).count()
        total_reports_all = PatientMaster.objects.filter(status=True).count()
        # ✅ Add search functionality
        query = request.GET.get("q", "")
        patients = PatientMaster.objects.filter(status=True).order_by('patient_name')
        if query:
            patients = patients.filter(
                patient_name__icontains=query
            ) | patients.filter(
                patient_id__icontains=query
            ) | patients.filter(
                mobile_number__icontains=query
            )

        # Prepare context with safe JSON serializable data
        context = {
            "patients": patients,
            "query": query,
            "test_count": test_count,
            "patient_count": patient_count,
            "doctor_count": doctor_count,
            "pending_reports": pending_reports,
            "total_reports": total_reports,
            "avg_daily_reports": avg_daily_reports,
            "total_reports_all": total_reports_all,
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "labels": json.dumps(date_labels),
            "daily_counts": json.dumps(daily_counts),
        }

        # ✅ AJAX refresh for charts
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'labels': date_labels,
                'daily_counts': daily_counts,
                'total_reports': total_reports,
                'avg_daily_reports': avg_daily_reports,
                'pending_reports': pending_reports
            })

        logger.debug(f"Dashboard data: {total_reports} reports, {patient_count} patients")
        # ✅ Keep your JS safe by keeping {"context": context}
        return render(request, "dashboard.html", {"context": context})
    except Exception as e:
        logger.error(f"Dashboard error: {str(e)}", exc_info=True)
        messages.error(request, "Error loading dashboard")
        return redirect('login')

def register_user(request):
    logger.info("User registration attempt")
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.username = user.username.lower()
                user.email = user.email.lower()
                user.save()
                logger.info(f"New user registered: {user.username}")
                messages.success(request, "Registration successful. Please login.")
                return redirect('login')
            except Exception as e:
                logger.error(f"Registration error: {str(e)}", exc_info=True)
                messages.error(request, f"Error: {str(e)}")
        else:
            logger.warning(f"Invalid registration form: {form.errors}")
            # Show validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        form = RegisterForm()

    return render(request, 'accounts/register.html', {'form': form})

def login_user(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data['username_or_email']
            password = form.cleaned_data['password']

            logger.info(f"Login attempt for user: {username_or_email}")
            # check login by username or email
            user = authenticate(request, username=username_or_email, password=password)
            if user is None:
                try:
                    user_obj = CustomUser.objects.get(email=username_or_email)
                    user = authenticate(request, username=user_obj.username, password=password)
                except CustomUser.DoesNotExist:
                    user = None

            if user is not None:
                login(request, user)
                logger.info(f"Successful login for user: {user.username}")
                messages.success(request, "Login successful.")  
                return redirect('dashboard')  # after login redirect to dashboard
            else:
                logger.warning(f"Failed login attempt for: {username_or_email}")
                messages.error(request, "Invalid username/email or password.")
    else:
        form = LoginForm()
    return render(request, 'accounts/login.html', {'form': form})

def logout_user(request):
    logger.info(f"User logged out: {request.user.username}")
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect('login')

def forgot_password(request):
    return render(request, 'accounts/forgot_password.html')

@login_required
def create_user(request):
    if request.method == 'POST':
        # Include request.FILES for image uploads
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "User created successfully.")
                return redirect('view_users')
            except Exception as e:
                messages.error(request, f"Error saving user: {e}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        form = CustomUserCreationForm()
    return render(request, 'accounts/create_user.html', {'form': form})
@login_required
def edit_user(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        # Include request.FILES here too
        form = CustomUserChangeForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "User updated successfully.")
                return redirect('view_users')
            except Exception as e:
                messages.error(request, f"Error updating user: {e}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        form = CustomUserChangeForm(instance=user)
    return render(request, 'accounts/edit_user.html', {'form': form, 'user': user})


@login_required(login_url='login')
def view_users(request):
    query = request.GET.get('q', '').strip()
    lab = request.user.lab
    if request.user.role =='admin' or request.user.is_superuser:
        users = CustomUser.objects.filter(status=True).order_by('id')
    else:  
        users = CustomUser.objects.filter(status=True, lab=lab).order_by('id')

    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    paginator = Paginator(users, 10)  # 5 users per page
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number)

    return render(request, 'accounts/view_users.html', {
        'users': users_page,
        'query': query,
    })

def delete_user(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        try:
            user.status = False  # Soft delete
            user.save()
            messages.success(request, "User deleted successfully.")
            return redirect('view_users')
        except Exception as e:
            messages.error(request, f"Error deleting user: {e}")
    return render(request, 'accounts/confirm_delete.html', {'user': user})

@login_required
def change_password(request):
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        user = request.user

        # Check current password
        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('dashboard')

        # Check if new passwords match
        if new_password != confirm_password:
            messages.error(request, 'New password and confirm password do not match.')
            return redirect('dashboard')

        # Check password length or strength (optional)
        if len(new_password) < 6:
            messages.error(request, 'Password must be at least 6 characters long.')
            return redirect('dashboard')

        # Change password
        user.set_password(new_password)
        user.save()

        # Keep user logged in after changing password
        update_session_auth_hash(request, user)

        messages.success(request, 'Your password has been changed successfully!')
        return redirect('dashboard')

    return render(request, 'accounts/change_password.html')



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from patient.models import PatientMaster
from doctor.models import Doctor
from datetime import datetime
from doctor.models import Doctor,DoctorLedger, DoctorTransaction

def patient_search(request):
    query = request.GET.get('q', request.GET.get('term', '')).strip()
    if not query:
        return JsonResponse({'patients': []})

    # Search in patient_id, patient_name, and mobile_number
    patients = PatientMaster.objects.filter(
        Q(patient_id__icontains=query) | 
        Q(patient_name__icontains=query) |
        Q(mobile_number__icontains=query)
    ).order_by('patient_name')[:50]  # Increased limit for better search experience
    
    # Check if this is an autocomplete request (from select2 or similar)
    if 'term' in request.GET:
        results = [{
            'id': p.patient_id, 
            'value': f"{p.patient_name} ({p.patient_id})",
            'patient_id': p.patient_id,
            'patient_name': p.patient_name,
            'mobile_number': p.mobile_number or ''
        } for p in patients]
        return JsonResponse(results, safe=False)
    
    # Regular search request
    results = [{
        'patient_id': p.patient_id,
        'patient_name': p.patient_name,
        'mobile_number': p.mobile_number or ''
    } for p in patients]
    
    return JsonResponse({'patients': results})


def doctor_search(request):
    query = request.GET.get('q', request.GET.get('term', '')).strip()
    if not query:
        return JsonResponse({'doctors': []})

    # Search in doctor fields
    doctors = Doctor.objects.filter(
        Q(doctor_id__icontains=query) | 
        Q(doctor_name__icontains=query) |
        Q(mobile_number__icontains=query) |
        Q(hospital_name__icontains=query) |
        Q(specialization__icontains=query)
    ).order_by('doctor_name')[:50]
    
    # Check if this is an autocomplete request
    if 'term' in request.GET:
        results = [{
            'id': d.id,
            'value': f"{d.doctor_name} ({d.specialization or 'No Specialization'})",
            'doctor_id': d.doctor_id,
            'doctor_name': d.doctor_name,
            'mobile_number': d.mobile_number or '',
            'specialization': d.specialization or '',
            'hospital_name': d.hospital_name or ''
        } for d in doctors]
        return JsonResponse(results, safe=False)
    
    # Regular search request
    results = [{
        'doctor_id': d.doctor_id,
        'doctor_name': d.doctor_name,
        'mobile_number': d.mobile_number or '',
        'specialization': d.specialization or '',
        'hospital_name': d.hospital_name or ''
    } for d in doctors]
    
    return JsonResponse({'doctors': results})


# ============================================================
# ✅ EXPORT DOCTORS TO EXCEL
# ============================================================
@login_required
def export_doctors_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Doctors"

    # Column headers
    headers = [
        'ID', 'Doctor Name', 'Hospital Name', 'Mobile Number',
        'Email', 'Address', 'Commission %', 'Status'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Fetch doctor data
    doctors = Doctor.objects.select_related('lab').filter(status=True).order_by('-id')

    for row_num, doctor in enumerate(doctors, 2):
        ws.cell(row=row_num, column=1, value=doctor.id)
        ws.cell(row=row_num, column=2, value=doctor.doctor_name)
        ws.cell(row=row_num, column=3, value=doctor.hospital_name)
        ws.cell(row=row_num, column=4, value=doctor.mobile_number)
        ws.cell(row=row_num, column=5, value=doctor.email)
        ws.cell(row=row_num, column=6, value=doctor.address)
        ws.cell(row=row_num, column=7, value=float(doctor.commission_percent) if doctor.commission_percent else 0.0)
        ws.cell(row=row_num, column=8, value='Active' if doctor.status else 'Inactive')

    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min((max_length + 2) * 1.2, 30)

    # Response setup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"doctors_export_{timestamp}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# ============================================================
# ✅ EXPORT SINGLE PATIENT TO EXCEL
# ============================================================


def export_patient_selection(request):
    patients = PatientMaster.objects.filter(status=True).order_by('patient_name')

    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        if patient_id:
            return redirect('export_patient_excel', patient_id=patient_id)

    return render(request, 'export_patient_selection.html', {'patients': patients})

@login_required
def export_patient_excel(request, patient_id):
    wb = Workbook()
    ws = wb.active

    try:
        # Fetch patient record
        patient = PatientMaster.objects.get(patient_id=patient_id, status=True)
        ws.title = f"Patient_{patient_id}"

        # Headers
        headers = [
            'Patient ID', 'Patient Name', 'Gender', 'Age', 'Mobile Number',
            'Address', 'Referred By', 'Sample Date', 'Reporting Date', 'Weight',
            'Test Names', 'Total Amount', 'Discount', 'Paid Amount', 'Balance Amount', 'Status'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Patient test data
        test_names = ", ".join([test.test.test_name for test in patient.tests.select_related('test').all()])
        ws.cell(row=2, column=1, value=patient.patient_id)
        ws.cell(row=2, column=2, value=str(patient))
        ws.cell(row=2, column=3, value=patient.gender)
        ws.cell(row=2, column=4, value=patient.age)
        ws.cell(row=2, column=5, value=patient.mobile_number)
        ws.cell(row=2, column=6, value=patient.address)
        ws.cell(row=2, column=7, value=str(patient.doctor) if patient.doctor else '')
        ws.cell(row=2, column=8, value=patient.sample_date.strftime('%Y-%m-%d') if patient.sample_date else '')
        ws.cell(row=2, column=9, value=patient.reporting_date.strftime('%Y-%m-%d') if patient.reporting_date else '')
        ws.cell(row=2, column=10, value=float(patient.weight) if patient.weight else 0.0)
        ws.cell(row=2, column=11, value=test_names)
        ws.cell(row=2, column=12, value=float(patient.total_amount) if patient.total_amount else 0.0)
        ws.cell(row=2, column=13, value=float(patient.discount) if patient.discount else 0.0)
        ws.cell(row=2, column=14, value=float(patient.paid_amount) if patient.paid_amount else 0.0)
        ws.cell(row=2, column=15, value=float(patient.balance_amount) if patient.balance_amount else 0.0)
        ws.cell(row=2, column=16, value='Active' if patient.status else 'Inactive')

        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min((max_length + 2) * 1.2, 30)

        # Response setup
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"patient_{patient_id}_report_{timestamp}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    except PatientMaster.DoesNotExist:
        return HttpResponse("Patient not found", status=404)

@login_required
def export_all_patients_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Patients"

    # Headers
    headers = [
        'Patient ID', 'Patient Name', 'Gender', 'Age', 'Mobile Number',
        'Address', 'Referred By', 'Sample Date', 'Reporting Date', 'Weight','test names',
        'Total Amount', 'Discount', 'Paid Amount', 'Balance Amount', 'Status'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Get all patients with their tests
    patients = PatientMaster.objects.filter(status=True).prefetch_related('tests__test')
    
    # Write patient data
    for row_num, patient in enumerate(patients, 2):
        # Get test names for the patient
        test_names = ", ".join([test.test.test_name for test in patient.tests.all()])
        
        ws.cell(row=row_num, column=1, value=patient.patient_id)
        ws.cell(row=row_num, column=2, value=str(patient))
        ws.cell(row=row_num, column=3, value=patient.gender)
        ws.cell(row=row_num, column=4, value=patient.age)
        ws.cell(row=row_num, column=5, value=patient.mobile_number)
        ws.cell(row=row_num, column=6, value=patient.address or '')
        ws.cell(row=row_num, column=7, value=str(patient.doctor) if patient.doctor else '')
        ws.cell(row=row_num, column=8, value=patient.sample_date.strftime('%Y-%m-%d') if patient.sample_date else '')
        ws.cell(row=row_num, column=9, value=patient.reporting_date.strftime('%Y-%m-%d') if patient.reporting_date else '')
        ws.cell(row=row_num, column=10, value=float(patient.weight) if patient.weight else 0.0)
        ws.cell(row=row_num, column=11, value=test_names)
        ws.cell(row=row_num, column=12, value=float(patient.total_amount) if patient.total_amount else 0.0)
        ws.cell(row=row_num, column=13, value=float(patient.discount) if patient.discount else 0.0)
        ws.cell(row=row_num, column=14, value=float(patient.paid_amount) if patient.paid_amount else 0.0)
        ws.cell(row=row_num, column=15, value=float(patient.balance_amount) if patient.balance_amount else 0.0)
        ws.cell(row=row_num, column=16, value='Active' if patient.status else 'Inactive')
   
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Cap at 30 characters

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_patients.xlsx'
    wb.save(response)
    return response


def backform(request):
    return render(request, 'Backupform.html')

def restoreform(request):
    return render(request, 'Restoreform.html')


from django.shortcuts import render, redirect
from django.contrib import messages
from pathlib import Path
import shutil
from datetime import datetime


def take_backup(request):
    if request.method == "POST":
        try:
            logger.info("Database backup initiated")
            BASE_DIR = Path(__file__).resolve().parent.parent
            DB_FILE = BASE_DIR / 'db.sqlite3'
            GOOGLE_DRIVE_FOLDER = Path(r"G:\My Drive\Pathsoft_backup")
            GOOGLE_DRIVE_FOLDER.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
            backup_filename = f"db_backup_{timestamp}.sqlite3"
            backup_path = GOOGLE_DRIVE_FOLDER / backup_filename

            if DB_FILE.exists():
                shutil.copy2(DB_FILE, backup_path)
                logger.info(f"Backup created successfully: {backup_filename}")
                # Keep only the 30 latest backups
                backups = sorted(GOOGLE_DRIVE_FOLDER.glob("db_backup_*.sqlite3"), reverse=True)
                for old_file in backups[30:]:
                    old_file.unlink()
                
                messages.success(request, f"Backup created successfully: {backup_filename}")
            else:
                logger.error("Database file not found for backup")
                messages.error(request, "Database file not found for backup.")
        except Exception as e:
            logger.error(f"Backup failed: {str(e)}", exc_info=True)
            messages.error(request, f"Backup failed: {e}")
    return redirect(request.META.get('HTTP_REFERER', '/'))


def restore_from_upload(request):
    """Replace old db.sqlite3 with uploaded backup file"""
    if request.method == "POST" and request.FILES.get("dbfile"):
        try:
            dbfile = request.FILES["dbfile"]
            BASE_DIR = Path(__file__).resolve().parent.parent
            DB_FILE = BASE_DIR / "db.sqlite3"
            BACKUP_DIR = BASE_DIR / "manual_backups"
            BACKUP_DIR.mkdir(exist_ok=True)

            # Backup current DB before replacing
            shutil.copy2(DB_FILE, BACKUP_DIR / f"db_backup_before_restore.sqlite3")

            # Save uploaded file temporarily
            temp_path = BACKUP_DIR / "uploaded_backup.sqlite3"
            with open(temp_path, "wb+") as destination:
                for chunk in dbfile.chunks():
                    destination.write(chunk)

            # Replace main DB with uploaded file
            shutil.copy2(temp_path, DB_FILE)
            messages.success(request, "Database successfully restored from uploaded file.")

        except Exception as e:
            messages.error(request, f"Restore failed: {e}")

    return redirect(request.META.get("HTTP_REFERER", "/"))


def restore_backup(request):
    """Restore latest DB backup from Google Drive folder and restart server."""
    if request.method == "POST":
        try:
            logger.info("Database restore initiated")
            BASE_DIR = Path(__file__).resolve().parent.parent
            DB_FILE = BASE_DIR / 'db.sqlite3'
            GOOGLE_DRIVE_FOLDER = Path(r"G:\My Drive\Pathsoft_backup")

            backups = sorted(GOOGLE_DRIVE_FOLDER.glob("db_backup_*.sqlite3"), reverse=True)
            if backups:
                latest_backup = backups[0]
                shutil.copy2(latest_backup, DB_FILE)
                logger.info(f"Database restored from: {latest_backup.name}")
                messages.success(request, f"Restored from backup: {latest_backup.name}")
            else:
                logger.warning("No backup files found for restore")
                messages.error(request, "No backup files found in Google Drive folder.")
        except Exception as e:
            logger.error(f"Restore failed: {str(e)}", exc_info=True)
            messages.error(request, f"Restore failed: {e}")
    return redirect(request.META.get('HTTP_REFERER', '/'))

# from pydrive.auth import GoogleAuth
# from pydrive.drive import GoogleDrive

# def take_backup(request):
#     """Take DB backup and upload to Google Drive using PyDrive."""
#     if request.method == "POST":
#         try:
#             BASE_DIR = Path(__file__).resolve().parent.parent
#             DB_FILE = BASE_DIR / 'db.sqlite3'

#             if not DB_FILE.exists():
#                 messages.error(request, "⚠️ Database file not found for backup.")
#                 return redirect(request.META.get('HTTP_REFERER', '/'))

#             # Prepare backup file
#             timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
#             backup_filename = f"db_backup_{timestamp}.sqlite3"
#             temp_backup_path = BASE_DIR / backup_filename
#             shutil.copy2(DB_FILE, temp_backup_path)

#             # Authenticate and upload using PyDrive
#             gauth = GoogleAuth()
#             gauth.LoadCredentialsFile("mycreds.txt")

#             if gauth.credentials is None:
#                 # First time authorization
#                 gauth.LocalWebserverAuth(port_numbers=[8020])
#             elif gauth.access_token_expired:
#                 gauth.Refresh()
#             else:
#                 gauth.Authorize()

#             # Save credentials for reuse
#             gauth.SaveCredentialsFile("mycreds.txt")

#             drive = GoogleDrive(gauth)

#             # Upload file to your Drive folder
#             folder_id = "1GzP9pAzVXbhePJUaSTBhbRzSOFBddkt1"  # Replace with your folder ID
#             file_drive = drive.CreateFile({'title': backup_filename, 'parents': [{'id': folder_id}]})
#             file_drive.SetContentFile(str(temp_backup_path))
#             file_drive.Upload()

#             # Cleanup local copy
#             temp_backup_path.unlink(missing_ok=True)

#             messages.success(request, f"✅ Backup uploaded successfully to Google Drive: {backup_filename}")

#         except Exception as e:
#             messages.error(request, f"❌ Backup failed: {e}")

#     return redirect(request.META.get('HTTP_REFERER', '/')) 

# from utils.google_drive_backup import upload_backup_to_drive

# def take_backup(request):
#     """Take manual DB backup and upload to Google Drive folder via API."""
#     if request.method == "POST":
#         try:
#             BASE_DIR = Path(__file__).resolve().parent.parent
#             DB_FILE = BASE_DIR / 'db.sqlite3'

#             timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
#             backup_filename = f"db_backup_{timestamp}.sqlite3"
#             temp_backup_path = BASE_DIR / backup_filename

#             if DB_FILE.exists():
#                 shutil.copy2(DB_FILE, temp_backup_path)
                
#                 # Upload to Google Drive
#                 drive_file_id = upload_backup_to_drive(temp_backup_path, backup_filename)

#                 messages.success(request, f"✅ Backup uploaded to Google Drive (File ID: {drive_file_id})")
                
#                 # Optionally delete local temp backup after upload
#                 temp_backup_path.unlink(missing_ok=True)
#             else:
#                 messages.error(request, "⚠️ Database file not found for backup.")
#         except Exception as e:
#             messages.error(request, f"❌ Backup failed: {e}")

#     return redirect(request.META.get('HTTP_REFERER', '/'))

